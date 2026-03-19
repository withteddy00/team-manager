from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from typing import Optional
from datetime import date
from io import BytesIO
from app.database import get_db
from app.models.models import Holiday, HolidayPayment, EgyptDuty, EgyptBeneficiary, User
from app.services.auth import get_current_user

router = APIRouter(prefix="/api/exports", tags=["Exports"])


def _gather_export_data(
    db: Session,
    event_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
):
    rows = []

    # Holidays
    if event_type is None or event_type == "holiday":
        hq = db.query(Holiday)
        if year:
            hq = hq.filter(extract("year", Holiday.date) == year)
        if month:
            hq = hq.filter(extract("month", Holiday.date) == month)
        if date_from:
            hq = hq.filter(Holiday.date >= date.fromisoformat(date_from))
        if date_to:
            hq = hq.filter(Holiday.date <= date.fromisoformat(date_to))

        holidays = hq.order_by(Holiday.date).all()
        for h in holidays:
            if h.worked and h.payments:
                for p in h.payments:
                    status_str = "Travaillé" if h.worked else ("Non travaillé" if h.worked is False else "En attente")
                    rows.append({
                        "date": h.date.isoformat(),
                        "type": "Jour Férié",
                        "nom": h.holiday_name,
                        "membre": p.member.full_name if p.member else "",
                        "montant": f"{p.amount:.0f} DH",
                        "statut": status_str,
                        "commentaire": h.comment or "",
                    })
            else:
                status_str = "Non travaillé" if h.worked is False else "En attente"
                rows.append({
                    "date": h.date.isoformat(),
                    "type": "Jour Férié",
                    "nom": h.holiday_name,
                    "membre": "-",
                    "montant": "0 DH",
                    "statut": status_str,
                    "commentaire": h.comment or "",
                })

    # Egypt duties
    if event_type is None or event_type == "egypt_duty":
        eq = db.query(EgyptDuty)
        if year:
            eq = eq.filter(extract("year", EgyptDuty.date) == year)
        if month:
            eq = eq.filter(extract("month", EgyptDuty.date) == month)
        if date_from:
            eq = eq.filter(EgyptDuty.date >= date.fromisoformat(date_from))
        if date_to:
            eq = eq.filter(EgyptDuty.date <= date.fromisoformat(date_to))

        duties = eq.order_by(EgyptDuty.date).all()
        for d in duties:
            for b in d.beneficiaries:
                rows.append({
                    "date": d.date.isoformat(),
                    "type": "Astreinte Égypte",
                    "nom": "Astreinte Dimanche",
                    "membre": b.member.full_name if b.member else "",
                    "montant": f"{b.amount:.0f} DH",
                    "statut": "Validé",
                    "commentaire": d.comment or "",
                })

    rows.sort(key=lambda x: x["date"], reverse=True)
    return rows


@router.get("/excel")
def export_excel(
    event_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = _gather_export_data(db, event_type, date_from, date_to, month, year)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"

    headers = ["Date", "Type", "Nom", "Membre", "Montant", "Statut", "Commentaire"]
    header_fill = PatternFill(start_color="1DB954", end_color="1DB954", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(["date", "type", "nom", "membre", "montant", "statut", "commentaire"], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historique.xlsx"},
    )


@router.get("/pdf")
def export_pdf(
    event_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    rows = _gather_export_data(db, event_type, date_from, date_to, month, year)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("Historique des Paiements - Team Manager", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 20))

    headers = ["Date", "Type", "Nom", "Membre", "Montant", "Statut", "Commentaire"]
    table_data = [headers]

    for row in rows:
        table_data.append([
            row["date"], row["type"], row["nom"],
            row["membre"], row["montant"], row["statut"], row["commentaire"]
        ])

    if len(table_data) == 1:
        table_data.append(["Aucune donnée", "", "", "", "", "", ""])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1DB954")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=historique.pdf"},
    )
